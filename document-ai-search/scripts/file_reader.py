#!/usr/bin/env python3
"""
Unified document reader for multiple file formats.

Supports: TXT, PDF, DOCX, DOC, XLSX, XLS, JPG, PNG, JPEG (OCR)

See SKILL.md for usage.
"""

import subprocess
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional


class ReadError(Exception):
    """Document read error with details."""
    pass


class DocumentReader:
    """Reads text from various document formats."""

    # Max characters to read per document (to avoid memory issues)
    MAX_CHARS = 100000

    def read(self, file: Path) -> str:
        """
        Read document based on file extension.

        Args:
            file: Path to the document file

        Returns:
            Extracted text content

        Raises:
            ReadError: If file cannot be read
        """
        ext = file.suffix.lower()

        readers = {
            '.txt': self._read_txt,
            '.pdf': self._read_pdf,
            '.docx': self._read_docx,
            '.doc': self._read_doc,   # Legacy Word format
            '.xlsx': self._read_xlsx,
            '.xls': self._read_xls,   # Legacy Excel format
            '.jpg': self._read_image, # OCR support
            '.jpeg': self._read_image,
            '.png': self._read_image,
            '.md': self._read_txt,    # Markdown as plain text
        }

        reader = readers.get(ext)
        if not reader:
            raise ReadError(f"Unsupported format: {ext}")

        try:
            text = reader(file)
            # Truncate if too long
            if len(text) > self.MAX_CHARS:
                text = text[:self.MAX_CHARS] + "\n\n[Content truncated...]"
            return text
        except ReadError:
            raise
        except Exception as e:
            raise ReadError(f"{file.name}: {type(e).__name__}: {e}")

    def _read_txt(self, file: Path) -> str:
        """Read plain text file with encoding fallback."""
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']

        for encoding in encodings:
            try:
                return file.read_text(encoding=encoding)
            except (UnicodeDecodeError, LookupError):
                continue

        raise ReadError("Could not decode file with any encoding")

    def _read_pdf(self, file: Path) -> str:
        """Extract text from PDF using pypdf."""
        try:
            from pypdf import PdfReader
        except ImportError:
            raise ReadError("pypdf not installed. Run: pip install pypdf")

        try:
            reader = PdfReader(str(file))

            # Check if encrypted
            if reader.is_encrypted:
                try:
                    reader.decrypt('')
                except Exception:
                    raise ReadError("Encrypted PDF (password protected)")

            text_parts = []
            for page_num, page in enumerate(reader.pages, 1):
                try:
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text_parts.append(page_text)
                except Exception as e:
                    # Skip problematic pages but continue
                    text_parts.append(f"[Page {page_num} error: {e}]")

            text = "\n\n".join(text_parts)

            if not text.strip():
                raise ReadError("No text content found (may be image-only PDF)")

            return text

        except ReadError:
            raise
        except Exception as e:
            raise ReadError(f"PDF read failed: {e}")

    def _read_docx(self, file: Path) -> str:
        """Extract text from DOCX by reading XML."""
        try:
            with zipfile.ZipFile(file, 'r') as docx:
                # Check if main document exists
                try:
                    xml_content = docx.read('word/document.xml')
                except KeyError:
                    raise ReadError("Invalid DOCX structure (missing word/document.xml)")

                # Parse XML with defusedxml for security
                try:
                    from defusedxml import ElementTree as SafeET
                    tree = SafeET.fromstring(xml_content)
                except ImportError:
                    # Fallback to standard library with warning
                    tree = ET.fromstring(xml_content)

                # Extract text from all w:t elements
                namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                texts = tree.findall('.//w:t', namespaces)

                result = []
                for t in texts:
                    if t.text:
                        result.append(t.text)

                text = ''.join(result)

                if not text.strip():
                    raise ReadError("No text content found")

                return text

        except ReadError:
            raise
        except Exception as e:
            raise ReadError(f"DOCX read failed: {e}")

    def _read_xlsx(self, file: Path) -> str:
        """Extract text from XLSX using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ReadError("openpyxl not installed. Run: pip install openpyxl")

        try:
            wb = load_workbook(str(file), read_only=True, data_only=True)
            sheets_text = []

            for sheet in wb:
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out empty rows
                    non_empty_cells = [str(cell) if cell is not None else '' for cell in row if cell is not None]
                    if non_empty_cells:
                        row_text = ' | '.join(non_empty_cells)
                        if row_text.strip():
                            rows_text.append(row_text)

                if rows_text:
                    sheets_text.append(f"Sheet: {sheet.title}\n" + "\n".join(rows_text))

            wb.close()

            if not sheets_text:
                raise ReadError("No data found in workbook")

            return "\n\n".join(sheets_text)

        except ReadError:
            raise
        except Exception as e:
            raise ReadError(f"XLSX read failed: {e}")

    def _read_xls(self, file: Path) -> str:
        """Extract text from legacy XLS using xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ReadError("xlrd not installed. Run: pip install xlrd")

        try:
            # Open workbook with xlrd
            workbook = xlrd.open_workbook(str(file), formatting_info=False, on_demand=True)
            sheets_text = []

            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)

                rows_text = []
                for row_idx in range(sheet.nrows):
                    # Get all cell values in this row
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        try:
                            cell_value = sheet.cell_value(row_idx, col_idx)
                            # Convert to string, handling different types
                            if cell_value == '':
                                continue
                            if isinstance(cell_value, float):
                                # Check if it's an integer (e.g., 42.0 -> 42)
                                if cell_value.is_integer():
                                    cell_value = str(int(cell_value))
                                else:
                                    cell_value = str(cell_value)
                            else:
                                cell_value = str(cell_value).strip()

                            if cell_value:
                                row_values.append(cell_value)
                        except Exception:
                            continue

                    if row_values:
                        row_text = ' | '.join(row_values)
                        rows_text.append(row_text)

                if rows_text:
                    sheets_text.append(f"Sheet: {sheet.name}\n" + "\n".join(rows_text))

                # Release sheet memory
                workbook.unload_sheet(sheet_idx)

            if not sheets_text:
                raise ReadError("No data found in workbook")

            return "\n\n".join(sheets_text)

        except ReadError:
            raise
        except Exception as e:
            raise ReadError(f"XLS read failed: {e}")

    def _read_doc(self, file: Path) -> str:
        """Extract text from legacy DOC using multiple methods."""
        # Method 1: Try pywin32 COM automation (Windows with Word installed)
        if sys.platform == "win32":
            try:
                import win32com.client
                word = win32com.client.Dispatch('Word.Application')
                word.Visible = False
                doc = word.Documents.Open(str(file))
                text = doc.Content.Text
                doc.Close(False)
                word.Quit()
                return text
            except ImportError:
                pass
            except Exception:
                pass

        # Method 2: Try antiword (Linux/Mac)
        if sys.platform != "win32":
            for tool in ['antiword', 'catdoc']:
                try:
                    result = subprocess.run(
                        [tool, str(file)],
                        capture_output=True,
                        text=True,
                        timeout=30
                    )
                    if result.returncode == 0 and result.stdout.strip():
                        return result.stdout
                except (FileNotFoundError, subprocess.TimeoutExpired):
                    continue

        # Method 3: Try LibreOffice (cross-platform)
        for libreoffice_cmd in ['libreoffice', 'soffice']:
            try:
                # Convert to temp text file
                result = subprocess.run(
                    [libreoffice_cmd, '--headless', '--convert-to', 'txt:Text',
                     '--outdir', str(file.parent), str(file)],
                    capture_output=True,
                    timeout=60
                )
                if result.returncode == 0:
                    temp_txt = file.with_suffix('.txt')
                    if temp_txt.exists():
                        text = temp_txt.read_text(encoding='utf-8', errors='ignore')
                        temp_txt.unlink()  # Clean up
                        return text
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue

        raise ReadError(
            "Could not read .doc file. Options:\n"
            "1. Install pywin32 (Windows with Word): pip install pywin32\n"
            "2. Install LibreOffice: https://www.libreoffice.org/\n"
            "3. Convert .doc to .docx for better compatibility"
        )

    def _read_image(self, file: Path) -> str:
        """Extract text from images using OCR (pytesseract)."""
        try:
            from PIL import Image
            import pytesseract
        except ImportError:
            raise ReadError("OCR dependencies not installed. Run: pip install pytesseract Pillow")

        try:
            # Configure Tesseract path for Windows
            if sys.platform == "win32":
                tess_paths = [
                    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                    r"C:\Tesseract-OCR\tesseract.exe",
                ]
                for path in tess_paths:
                    if Path(path).exists():
                        pytesseract.pytesseract.tesseract_cmd = path
                        break

            # Open image and extract text
            image = Image.open(file)

            # Try Chinese OCR first, fallback to English
            try:
                text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            except pytesseract.TesseractError:
                # Chinese language data not available, use English only
                text = pytesseract.image_to_string(image, lang='eng')

            if not text or not text.strip():
                raise ReadError("No text could be extracted from image")

            return text

        except ReadError:
            raise
        except Exception as e:
            # OCR failures should not crash the search
            # Return empty string with a note
            return f"[OCR failed: {e}]"


def main():
    """CLI for testing file reader."""
    import sys

    if len(sys.argv) < 2:
        print("Usage: python file_reader.py <file_path>")
        sys.exit(1)

    file_path = Path(sys.argv[1])

    if not file_path.exists():
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    reader = DocumentReader()

    try:
        content = reader.read(file_path)
        print(f"Successfully read {file_path.name}")
        print(f"Content length: {len(content)} characters")
        print("-" * 60)
        print(content[:2000])  # Show first 2000 chars
        if len(content) > 2000:
            print(f"\n... ({len(content) - 2000} more characters)")
    except ReadError as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
